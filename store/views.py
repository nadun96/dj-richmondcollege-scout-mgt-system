from django.db.models import Count
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import JsonResponse, HttpResponse
from django.db.models import Max
from .utility import generate_code
from django.db.models.functions import ExtractWeekDay
from django.shortcuts import render
from core.models import Profile
from store.models import Item, Broken, Lend
import datetime
from django.db import transaction
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from .forms import LendForm, ReturnLendForm
import pandas as pd
from core.models import Group

""" view main items screen """


@login_required()
def main(request):

    context = {
        'title': 'main',
    }

    return render(request, 'store/main', context)


""" view items as a table """


@login_required()
def items(request):

    items = Item.objects.all()
    context = {
        'title': 'items',
        'items': items,
    }
    return render(request, 'store/items', context)


"""  view lend items screen """
lend_form = LendForm()
return_form = ReturnLendForm()


@login_required()
def lend(request):
    items = Lend.objects.filter(item_is_lent=True)
    context = {
        'title': 'lend',
        'lends': lend_form,
        'returns': return_form,
        'items': items,
    }
    return render(request, 'store/lend', context)


""" view reports screen  """


@login_required()
def reports(request):
    lend_users = Lend.objects.all().select_related('users')
    lend_items = Lend.objects.all().select_related('item')
    context = {
        'title': 'reports',
        'lend_user': lend_users,
        'lend_items': lend_items,
    }

    return render(request, 'store/reports', context)


""" view broken items as a table  """


@ login_required()
def broken(request):
    brokens = Broken.objects.all().select_related('item')

    context = {
        'title': 'broken',
        'brokens': brokens,
    }

    return render(request, 'store/broken', context)


""" add a new item  """


@ login_required()
def add_item(request):
    try:
        if request.method == 'POST':

            # store post request variables
            print(request.POST)
            item = request.POST.get('item_name')
            qty = int(request.POST.get('item_qty'))
            unit = request.POST.get('item_unit')
            price = int(request.POST.get('item_price'))

            # if item exists update quantity else create new item
            max_id = Item.objects.all().aggregate(Max('id'))['id__max']
            next_id = max_id + 1
            next_code = generate_code(next_id)
            print(f"Next Code is : {next_code}")

            with transaction.atomic():
                it, created = Item.objects.get_or_create(
                    # check for similarity
                    item_name=item, item_price=price,
                    # set default values dont check for similarity
                    defaults={
                        'item_code': next_code,
                        'item_quantity_received': qty,
                        'item_quantity_available': qty,
                        'item_purchased_date': datetime.datetime.now(),
                        'item_units': unit
                    }
                )
                # check what happened in the get_or_create method

            # if created is false then item was updated
            if (created == False):
                with transaction.atomic():
                    it.item_quantity_received = it.item_quantity_received + \
                        int(qty)
                    it.item_quantity_available = it.item_quantity_available + \
                        int(qty)
                    it.save()
                    print("item updated")
                    context = {
                        'title': 'main',
                        'message': f'{it} updated',
                    }

            # if created is true then item was created
            else:
                print("item added")
                context = {
                    'title': 'item added',
                    'message': 'item added successfully',
                }
            # return to page
            return render(request, 'store/main', context)

    except Exception as e:
        print(e)
        context = {
            'title': 'main',
            'message': 'item was not added',
        }
        return render(request, 'store/main', context)


""" send items to broken list  """


@ login_required()
def add_broken(request):
    """ send items to broken list """

    context = {}
    try:
        if request.method == 'POST':
            # save request params
            item = request.POST.get('item_code')
            qty = int(request.POST.get('item_qty'))

            item = int(item.lstrip('0'))
            print(item)  # Output: 1

            # check if the item is already in the items table
            exist = Item.objects.filter(id=item).filter(
                Q(item_quantity_available__gte=qty)).exists()

            print(exist)

            if (exist == True):

                with transaction.atomic():
                    # minus available quantity from items table
                    it = Item.objects.get(id=item)
                    it.item_quantity_available = it.item_quantity_available - \
                        qty
                    it.save()
                    print(it)

                    # add record to broken table at the same time
                    br = Broken.objects.create(
                        item_id=item,
                        item_quantity_broken=qty,
                        item_is_broken=True,
                    )
                    br.save()
                    print(br)
                    context = {
                        'title': 'main',
                        'message': f'{qty} {it} marked as broken',
                    }

                    return render(request, 'store/main', context)

            else:
                context = {
                    'title': 'main',
                    'message': 'item does not exist',
                }

                return render(request, 'store/main', context)

    except Exception as e:
        print(e)
        context = {
            'title': 'broken',
            'message': 'item not added',
        }
        return render(request, 'store/main', context)


""" bring repaired items back to items list """


@ login_required()
def add_repaired(request):

    context = {

    }

    try:
        if request.method == 'POST':

            # save params
            item = request.POST.get('item_id')
            qty = int(request.POST.get('item_qty'))

            item = int(item.lstrip('0'))
            print(item)  # Output: 1

            # get current details from items existence
            exist_in_items = Item.objects.filter(id=item).exists()

            # exist in broken table id equals and quantity is greater than or equal to
            exist_in_broken = Broken.objects.filter(
                Q(item_id=item) & Q(item_quantity_broken__gte=qty)).exists()

            if (exist_in_items and exist_in_broken):
                with transaction.atomic():
                    # get latest record from broken table
                    br = Broken.objects.filter(
                        Q(item_id=item) & Q(item_quantity_broken__gte=qty)).latest('id')
                    print(br)
                    # get item details from items table
                    it = Item.objects.get(
                        id=item,
                    )
                    print(it)

                # get current details from broken table
                # br = Broken.objects.filter(item_id=item).latest('id')

                # check validity of data
                if (br.item_quantity_broken >= qty):

                    with transaction.atomic():

                        # remove from broken table
                        br.item_quantity_broken = br.item_quantity_broken - qty
                        br.item_is_broken = True
                        br.save()

                        # add to items table
                        it.item_quantity_available = it.item_quantity_available + \
                            int(qty)
                        it.save()

                        # set 0 or minus to is broken true
                        br = Broken.objects.filter(
                            Q(item_quantity_broken__lte=0)
                        ).all().update(item_is_broken=False)

                    context = {
                        'title': 'main',
                        'message': 'item returned successfully',
                    }

                    return render(request, 'store/main', context)

            else:
                # set 0 to is broken
                br = Broken.objects.filter(
                    item_quantity_broken__exact=0).all().update(item_is_broken=True)

                context = {
                    'title': 'main',
                    'message': 'item does not exist',
                }

                return render(request, 'store/main', context)
    except Exception as e:
        print(e)
        context = {
            'title': 'items',
            'message': 'an error occurred',
        }

        return render(request, 'store/main', context)


""" add items to the lend list  """


@ login_required()
def add_lend(request):
    context = {'result': 'not loaded'}
    if request.method == 'POST':
        user = request.POST.get('user')
        item = request.POST.get('item')
        item = int(item.lstrip('0'))
        qty = int(request.POST.get('item_quantity_lent'))

        print(item)  # Output: 1

        # get users list from users table
        exist_user = Profile.objects.filter(
            Q(id=user)).exists()

        # check availability of items
        exist_items = Item.objects.filter(
            Q(id=item) & Q(item_quantity_available__gte=qty)).exists()

        print(f'item - {exist_items}, user - {exist_user}')

        # if all ok add to lend table
        if (exist_user and exist_items):
            user = Profile.objects.get(id=user)
            item = Item.objects.get(id=item)

            # add to lend table
            ln = Lend.objects.create(
                user=user,
                item=item,
                item_quantity_lent=qty,
                # set is lent to true
                item_is_lent=True,
            )

            # reduce from item
            item.item_quantity_available = item.item_quantity_available = item.item_quantity_available - \
                qty

            with transaction.atomic():
                ln.save()
                item.save()
                context = {'result': 'success'}
        else:
            print('not valid')
            context = {'result': 'exceeded'}

    return HttpResponse(JsonResponse(context))


""" return items from lend list  """


@ login_required()
def return_lend(request):
    context = {'result': 'not loaded'}
    if request.method == 'POST':
        form = ReturnLendForm(request.POST)
        valid = form.is_valid()
        if (valid):
            lend = request.POST.get('lend')
            lend = Lend.objects.get(id=lend)
            item = Item.objects.get(id=lend.item_id)

            # check if item available is greater than quantity lent
            if (item.item_quantity_available < item.item_quantity_received):
                less = True

            if (less):
                lend.item_is_lent = False
                lend.date_returned_date = datetime.datetime.now()
                item.item_quantity_available = item.item_quantity_available\
                    + lend.item_quantity_lent

                with transaction.atomic():
                    item.save()
                    lend.save()
                    context = {'result': 'success'}

        else:
            print('not valid')
            context = {'result': 'fail'}

    return HttpResponse(JsonResponse(context))


# @ login_required()
# def return_lend(request):

#     items = Lend.objects.filter(item_is_lent=True)

#     try:
#         if request.method == 'POST':
#             user = int(request.POST.get('user'))
#             item = request.POST.get('item')

#             # item = int(item.lstrip('0'))
#             print(item)  # Output: 1

#             # for query
#             item = Item.objects.get(id=item)
#             user = Profile.objects.get(id=user)

#             # if exist get the latest record that matches item and quantity from lend table
#             exist_available_lend = Lend.objects.filter(
#                 Q(item=item) & Q(user=user) & Q(item_is_lent=True)).exists()
#             # print(exist_available_lend)

#             # check if item available is greater than quantity lent
#             if (item.item_quantity_available < item.item_quantity_received):
#                 less = True

#             # if available add to items to lend table
#             if (exist_available_lend == True and less == True):

#                 with transaction.atomic():

#                     # update lend table
#                     ln = Lend.objects.filter(item=item).filter(
#                         user=user).filter(item_is_lent=True)\
#                         .latest('id')

#                     ln.item_is_lent = False

#                     ln.date_returned_date = datetime.datetime.now()

#                     ln.save()

#                     # add back to items availability
#                     it = Item.objects.get(id=item.id)

#                     it.item_quantity_available = it.item_quantity_available\
#                         + ln.item_quantity_lent

#                     it.save()

#                 context = {
#                     'title': 'lend',
#                     'items': items,
#                     'lends': lend_form,
#                     'message': 'record added',
#                 }
#                 return render(request, 'store/lend', context)

#             else:
#                 context = {
#                     'title': 'lend',
#                     'items': items,
#                     'lends': lend_form,
#                     'message': 'record does not exist',
#                 }
#                 return render(request, 'store/lend', context)

#     except Exception as e:
#         print(e)
#         context = {
#             'title': 'lend',
#             'items': items,
#             'lends': lend_form,
#             'message': 'record failed',
#         }
#         return render(request, 'store/lend', context)


""" get items data from database  to  json"""


@ login_required()
def get_items(request):
    items = Item.objects.all().values()
    context = {'items': items}
    return HttpResponse(JsonResponse(context, safe=False))


""" get lends data from database  to  json """


@ login_required()
def get_lends(request):
    lends = Lend.objects.all().select_related('user', 'user__user', 'item')\
        .values(
            'id',
            'user__user__username',
            'user__surname',
            'user__patrol',
            'item__item_name',
            'item_lent_date',
            'item_quantity_lent',
            'item_is_lent'
    ).all()
    context = {'lends': lends}
    return HttpResponse(JsonResponse(context, safe=False))


""" grouped by patrol """


@ login_required()
def get_lends_patrol(request):
    lends = Lend.objects.all().select_related('user', 'user__user', 'item')\
        .values(
            'user__patrol'
    ).annotate(
            count=Count('user__patrol')
    ).group_by('user__patrol').order_by('user__patrol')

    context = {'lends': lends}
    return HttpResponse(JsonResponse(context, safe=False))


""" get broken items data from database  to  json """


@ login_required()
def get_broken(request):
    broken = Broken.objects.all().values()
    context = {'broken': broken}
    return HttpResponse(JsonResponse(context, safe=False))


def get_users_lends(request):
    # lends with users who took items and their patrol
    pass


""" export lend data to excel """


@ login_required()
def export_lend(request):
    # create a queryset of records to export
    records = Lend.objects.all().select_related('user', 'user__user', 'item')\
        .values(
            'user__user__username',
            'user__surname',
            'user__patrol',
            'item__item_name',
            'item_lent_date',
            'item_quantity_lent',
            'item_is_lent'
    ).all()

    # print(records)

    # create the Excel file
    wb = openpyxl.Workbook()

    ws = wb.active

    # write the field names to the first row
    ws.append([
        'username',
        'surname',
        'patrol',
        'item',
        'date',
        'quantity',
        'item_is_lent'
    ])

    # iterate over the records and write the data to the sheet
    for record in records:
        ws.append([
            record['user__user__username'],
            record['user__surname'],
            record['user__patrol'],
            record['item__item_name'],
            record['item_lent_date'],
            record['item_quantity_lent'],
            record['item_is_lent'],
        ])

    # create an HttpResponse object with the Excel file as an attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=lend data{datetime.datetime.now()} .xlsx'

    wb.save(response)

    return response


""" export items data to excel """


@ login_required()
def export_items(request):
    # create a queryset of records to export
    records = Item.objects.all().\
        values(
        'item_code',
        'item_name',
        'item_price',
        'item_quantity_received',
        'item_units',
        'item_quantity_available',
        'item_purchased_date'
    )

    print(records)

    # create the Excel file
    wb = openpyxl.Workbook()

    ws = wb.active

    # write the field names to the first row
    ws.append([
        'Code',
        'Name',
        'Price',
        'Received',
        'Units',
        'Available',
        'Date'
    ])

    # iterate over the records and write the data to the sheet
    for record in records:
        ws.append([
            record['item_code'],
            record['item_name'],
            record['item_price'],
            record['item_quantity_received'],
            record['item_units'],
            record['item_quantity_available'],
            record['item_purchased_date']

        ])

    # create an HttpResponse object with the Excel file as an attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=items data - {datetime.datetime.now()} .xlsx'
    wb.save(response)

    return response


""" export broken data to excel """


@login_required()
def export_broken(request):

    # create a queryset of records to export
    records = Broken.objects.all().select_related('item')\
        .values(
            'item__item_code',
            'item__item_name',
            'item_quantity_broken',
            'item_broken_date',
            'item_is_broken',
            'date_repaired',
    ).all()

    print(records)

    # create the Excel file
    wb = openpyxl.Workbook()

    ws = wb.active

    # write the field names to the first row
    ws.append([
        'Code',
        'Item Name',
        'Quantity broken',
        'Date',
        'Broken',
        'date_repaired',
    ])

    # iterate over the records and write the data to the sheet
    for record in records:
        ws.append([
            record['item__item_code'],
            record['item__item_name'],
            record['item_quantity_broken'],
            record['item_broken_date'],
            record['item_is_broken'],
            record['date_repaired'],
        ])

    # create an HttpResponse object with the Excel file as an attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=broken data{datetime.datetime.now()} .xlsx'
    wb.save(response)

    return response


"""Export items formatted"""


@login_required()
def export_items_form(request):

    # group info
    group = Group.objects.get(id=1)
    print(group)

    # create a queryset of records to export
    records = Item.objects.all().\
        values(
        'item_code',
        'item_name',
        'item_price',
        'item_quantity_received',
        'item_units',
        'item_quantity_available',
        'item_purchased_date'
    )

    workbook_name = f'items data - {datetime.datetime.now()}'
    # create the Excel file workbook_name,
    wb = openpyxl.Workbook(write_only=False)

    # set the active worksheet
    ws = wb.create_sheet(title="Items Report", index=0,)
    wb.active = ws

    # first cell in the first row organization
    ws['A1'].font = Font(size=36, bold=True, color="800000")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill("solid", fgColor="D3D3D3")
    ws['A1'] = group.name

    # write the title to first row the second cell reaport title
    ws['C1'].font = Font(size=16, bold=True, color="00008B")
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C1'].fill = PatternFill("solid", fgColor="ADD8E6")
    ws['C1'] = 'Items Report as at {}'.format(
        datetime.datetime.now().strftime("%Y-%m-%d"))
    ws.row_dimensions[1].height = 50
    ws.merge_cells("C1:G1")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[2].height = 20

    # 2nd row item cout
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['A2'].fill = PatternFill("solid", fgColor="D3D3D3")
    number_of_items = Item.objects.all().count()
    ws['A2'] = 'Total Number of Items: {}'.format(number_of_items)
    ws.merge_cells("A2:G2")

    # write the field names to the second row
    ws.append([
        'Code',
        'Name',
        'Price(Rs.)',
        'Received(Qty)',
        'Units',
        'Available(Qty)',
        'Date'
    ])

    ws.row_dimensions[3].height = 20

    # set the font and cell background color for the field names
    for row in ws.iter_rows(min_row=3, max_col=7):
        for j, cell in enumerate(row):
            cell.font = Font(size=12, bold=True)
            cell.fill = PatternFill("solid", fgColor="FFA500")
            if j >= 2:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center")

    current_max_row = ws.max_row+1

    # iterate over the records and write the data to the sheet
    for i, record in enumerate(records):
        ws.append([
            record['item_code'],
            record['item_name'],
            record['item_price'],
            record['item_quantity_received'],
            record['item_units'],
            record['item_quantity_available'],
            record['item_purchased_date']

        ])
        # set the cell background color for alternating rows
        if i % 2 == 0:
            for row in ws.iter_rows(min_row=i+current_max_row, max_col=7):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="f2f2f2")
        else:
            for row in ws.iter_rows(min_row=i+current_max_row, max_col=7):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="ffffff")

    # current max row
    current_max_row = ws.max_row+1

    # set the column widths
    for i, column in enumerate(ws.columns):
        column_width = max(len(str(cell.value))
                           for cell in column[2:]) + 5
        ws.column_dimensions[openpyxl.utils.get_column_letter(
            i + 1)].width = column_width

    # add the filter command
    ws.auto_filter.ref = "A3:G3"

    # merge row before footer
    merge_start_row = ws.max_row + 1
    ws.cell(row=merge_start_row, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.merge_cells(start_row=merge_start_row, start_column=1,
                   end_row=merge_start_row, end_column=7)
    ws.cell(row=merge_start_row, column=1)

    # increase height of merged cells
    ws.row_dimensions[merge_start_row].height = 20

    # colour merged cells
    # merged_cell = ws.cell(row=merge_start_row, column=1)
    # merged_cell.fill = PatternFill("solid", fgColor="ffffff")

    # write footer
    footer_row = ws.max_row + 1

    # footer row 1
    # 1st column
    ws.cell(row=footer_row, column=1, value="Telephone").font = Font(italic=True,
                                                                     size=11, bold=True)
    ws.cell(row=footer_row, column=1).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=1).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=1,
                   end_row=footer_row, end_column=2)
    # 2nd column
    ws.cell(row=footer_row, column=3, value="Address").font = Font(italic=True,
                                                                   size=11, bold=True)
    ws.cell(row=footer_row, column=3).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=3).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=3,
                   end_row=footer_row, end_column=5)
    # 3rd column
    ws.cell(row=footer_row, column=6, value="Email").font = Font(italic=True,
                                                                 size=11, bold=True)
    ws.cell(row=footer_row, column=6).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=6).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=6,
                   end_row=footer_row, end_column=7)

    ws.row_dimensions[footer_row].height = 20

    # footer row 2
    # 1st column
    ws.cell(row=footer_row + 1, column=1,
            value=group.telephone).font = Font(size=11)
    ws.cell(row=footer_row+1, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=1).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=1,
                   end_row=footer_row + 1, end_column=2)
    # 2nd column
    ws.cell(row=footer_row + 1, column=3,
            value=group.address).font = Font(size=11)
    ws.cell(row=footer_row+1, column=3).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=3).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=3,
                   end_row=footer_row + 1, end_column=5)

    # 3rd column
    ws.cell(row=footer_row + 1, column=6,
            value=group.email).font = Font(size=11)
    ws.cell(row=footer_row+1, column=6).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=6).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=6,
                   end_row=footer_row + 1, end_column=7)

    ws.row_dimensions[footer_row+1].height = 50

    # footer row 3 system name
    system_name = f"Report was generated by RCSG MIS on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.cell(row=footer_row + 2, column=1,
            value=system_name).font = Font(italic=True, color="800000", size=10)
    ws.cell(row=footer_row + 2, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 2, column=1).alignment = Alignment(wrap_text=True,
                                                                horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 2, start_column=1,
                   end_column=7, end_row=footer_row + 2)

    # footer row 4 copyrights
    ws.cell(row=footer_row + 3, column=1,
            value="© 2023 RCSG MIS developers").font = Font(italic=True, color="800000", size=10)
    ws.cell(row=footer_row + 3, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 3, column=1).alignment = Alignment(wrap_text=True,
                                                                horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 3, start_column=1,
                   end_column=7, end_row=footer_row + 3)

    # set print area
    ws.print_area = "A1:G" + str(ws.max_row)

    # set page orientation to landscape

    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # create an HttpResponse object with the Excel file as an attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=items data - {dt} .xlsx'
    wb.save(response)

    return response


"""Export lends formatted"""


@login_required()
def export_lends_form(request):

    # group info
    group = Group.objects.get(id=1)
    print(group)

    # create a queryset of records to export
    records = Lend.objects.all().select_related('user', 'user__patrol', 'user__user', 'item')\
        .values(
            'user__user__username',
            'user__surname',
            'user__patrol__name',
            'item__item_name',
            'item_lent_date',
            'item_quantity_lent',
            'item_is_lent'
    ).all()

    workbook_name = f'items data - {datetime.datetime.now()}'
    # create the Excel file workbook_name,
    wb = openpyxl.Workbook(write_only=False)

    # set the active worksheet
    ws = wb.create_sheet(title="Lends Report", index=0,)
    wb.active = ws

    # first cell in the first row organization
    ws['A1'].font = Font(size=36, bold=True, color="800000")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill("solid", fgColor="D3D3D3")
    ws['A1'] = group.name

    # write the title to first row the second cell reaport title
    ws['C1'].font = Font(size=16, bold=True, color="00008B")
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C1'].fill = PatternFill("solid", fgColor="ADD8E6")
    ws['C1'] = 'Lends Report as at {}'.format(
        datetime.datetime.now().strftime("%Y-%m-%d"))
    ws.row_dimensions[1].height = 50
    ws.merge_cells("C1:G1")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[2].height = 20

    # 2nd row item cout
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['A2'].fill = PatternFill("solid", fgColor="D3D3D3")
    number_of_lends = Lend.objects.all().count()
    ws['A2'] = 'Total Number of Lends: {}'.format(number_of_lends)
    ws.merge_cells("A2:G2")

    # write the field names to the second row
    ws.append([
        'Username',
        'Surname',
        'Patrol',
        'Item',
        'Date',
        'Quantity',
        'Not Returned'
    ])

    ws.row_dimensions[3].height = 20

    # set the font and cell background color for the field names
    for row in ws.iter_rows(min_row=3, max_col=7):
        for j, cell in enumerate(row):
            cell.font = Font(size=12, bold=True)
            cell.fill = PatternFill("solid", fgColor="FFA500")
            if j >= 2:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center")

    current_max_row = ws.max_row+1

    # iterate over the records and write the data to the sheet
    for i, record in enumerate(records):
        is_lent = record['item_is_lent']
        if is_lent:
            is_lent_text = "Yes"
            fill_color = "ff9999"
        else:
            is_lent_text = "No"
        ws.append([
            record['user__user__username'],
            record['user__surname'],
            record['user__patrol__name'],
            record['item__item_name'],
            record['item_lent_date'],
            record['item_quantity_lent'],
            is_lent_text,
        ])
        # set the cell background color for alternating rows
        if i % 2 == 0:
            for row in ws.iter_rows(min_row=i+current_max_row, max_col=7):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="f2f2f2")
        else:
            for row in ws.iter_rows(min_row=i+current_max_row, max_col=7):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="ffffff")
        if is_lent:
            for row in ws.iter_rows(min_row=i+current_max_row, max_col=7):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)

    # current max row
    current_max_row = ws.max_row+1

    # set the column widths
    for i, column in enumerate(ws.columns):
        column_width = max(len(str(cell.value))
                           for cell in column[2:]) + 5
        ws.column_dimensions[openpyxl.utils.get_column_letter(
            i + 1)].width = column_width

    # add the filter command
    ws.auto_filter.ref = "A3:G3"

    # merge row before footer
    merge_start_row = ws.max_row + 1
    ws.cell(row=merge_start_row, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.merge_cells(start_row=merge_start_row, start_column=1,
                   end_row=merge_start_row, end_column=7)
    ws.cell(row=merge_start_row, column=1)

    # increase height of merged cells
    ws.row_dimensions[merge_start_row].height = 20

    # write footer
    footer_row = ws.max_row + 1

    # footer row 1
    # 1st column
    ws.cell(row=footer_row, column=1, value="Telephone").font = Font(italic=True,
                                                                     size=11, bold=True)
    ws.cell(row=footer_row, column=1).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=1).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=1,
                   end_row=footer_row, end_column=2)
    # 2nd column
    ws.cell(row=footer_row, column=3, value="Address").font = Font(italic=True,
                                                                   size=11, bold=True)
    ws.cell(row=footer_row, column=3).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=3).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=3,
                   end_row=footer_row, end_column=5)
    # 3rd column
    ws.cell(row=footer_row, column=6, value="Email").font = Font(italic=True,
                                                                 size=11, bold=True)
    ws.cell(row=footer_row, column=6).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=6).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=6,
                   end_row=footer_row, end_column=7)

    ws.row_dimensions[footer_row].height = 20

    # footer row 2
    # 1st column
    ws.cell(row=footer_row + 1, column=1,
            value=group.telephone).font = Font(size=11)
    ws.cell(row=footer_row+1, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=1).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=1,
                   end_row=footer_row + 1, end_column=2)
    # 2nd column
    ws.cell(row=footer_row + 1, column=3,
            value=group.address).font = Font(size=11)
    ws.cell(row=footer_row+1, column=3).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=3).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=3,
                   end_row=footer_row + 1, end_column=5)

    # 3rd column
    ws.cell(row=footer_row + 1, column=6,
            value=group.email).font = Font(size=11)
    ws.cell(row=footer_row+1, column=6).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=6).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=6,
                   end_row=footer_row + 1, end_column=7)

    ws.row_dimensions[footer_row+1].height = 50

    # footer row 3 system name
    system_name = f"Report was generated by RCSG MIS on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.cell(row=footer_row + 2, column=1,
            value=system_name).font = Font(italic=True, color="800000", size=10)
    ws.cell(row=footer_row + 2, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 2, column=1).alignment = Alignment(wrap_text=True,
                                                                horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 2, start_column=1,
                   end_column=7, end_row=footer_row + 2)

    # footer row 4 copyrights
    ws.cell(row=footer_row + 3, column=1,
            value="© 2023 RCSG MIS developers").font = Font(italic=True, color="800000", size=10)
    ws.cell(row=footer_row + 3, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 3, column=1).alignment = Alignment(wrap_text=True,
                                                                horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 3, start_column=1,
                   end_column=7, end_row=footer_row + 3)

    # set print area
    ws.print_area = "A1:G" + str(ws.max_row)

    # set page orientation to landscape

    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # create an HttpResponse object with the Excel file as an attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=items data - {dt} .xlsx'
    wb.save(response)

    return response


""" Export Broken Items Details  -- not working -- not prepared yet"""


@login_required()
def export_broken_form(request):

    # group info
    group = Group.objects.get(id=1)
    print(group)

    # create a queryset of records to export
    records = Lend.objects.all().select_related('user', 'user__patrol', 'user__user', 'item')\
        .values(
            'user__user__username',
            'user__surname',
            'user__patrol__name',
            'item__item_name',
            'item_lent_date',
            'item_quantity_lent',
            'item_is_lent'
    ).all()

    workbook_name = f'items data - {datetime.datetime.now()}'
    # create the Excel file workbook_name,
    wb = openpyxl.Workbook(write_only=False)

    # set the active worksheet
    ws = wb.create_sheet(title="Lends Report", index=0,)
    wb.active = ws

    # first cell in the first row organization
    ws['A1'].font = Font(size=36, bold=True, color="800000")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill("solid", fgColor="D3D3D3")
    ws['A1'] = group.name

    # write the title to first row the second cell reaport title
    ws['C1'].font = Font(size=16, bold=True, color="00008B")
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C1'].fill = PatternFill("solid", fgColor="ADD8E6")
    ws['C1'] = 'Lends Report as at {}'.format(
        datetime.datetime.now().strftime("%Y-%m-%d"))
    ws.row_dimensions[1].height = 50
    ws.merge_cells("C1:G1")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[2].height = 20

    # 2nd row item cout
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['A2'].fill = PatternFill("solid", fgColor="D3D3D3")
    number_of_lends = Lend.objects.all().count()
    ws['A2'] = 'Total Number of Lends: {}'.format(number_of_lends)
    ws.merge_cells("A2:G2")

    # write the field names to the second row
    ws.append([
        'Username',
        'Surname',
        'Patrol',
        'Item',
        'Date',
        'Quantity',
        'Not Returned'
    ])

    ws.row_dimensions[3].height = 20

    # set the font and cell background color for the field names
    for row in ws.iter_rows(min_row=3, max_col=7):
        for j, cell in enumerate(row):
            cell.font = Font(size=12, bold=True)
            cell.fill = PatternFill("solid", fgColor="FFA500")
            if j >= 2:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center")

    current_max_row = ws.max_row+1

    # iterate over the records and write the data to the sheet
    for i, record in enumerate(records):
        is_lent = record['item_is_lent']
        if is_lent:
            is_lent_text = "Yes"
            fill_color = "ff9999"
        else:
            is_lent_text = "No"
        ws.append([
            record['user__user__username'],
            record['user__surname'],
            record['user__patrol__name'],
            record['item__item_name'],
            record['item_lent_date'],
            record['item_quantity_lent'],
            is_lent_text,
        ])
        # set the cell background color for alternating rows
        if i % 2 == 0:
            for row in ws.iter_rows(min_row=i+current_max_row, max_col=7):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="f2f2f2")
        else:
            for row in ws.iter_rows(min_row=i+current_max_row, max_col=7):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="ffffff")
        if is_lent:
            for row in ws.iter_rows(min_row=i+current_max_row, max_col=7):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)

    # current max row
    current_max_row = ws.max_row+1

    # set the column widths
    for i, column in enumerate(ws.columns):
        column_width = max(len(str(cell.value))
                           for cell in column[2:]) + 5
        ws.column_dimensions[openpyxl.utils.get_column_letter(
            i + 1)].width = column_width

    # add the filter command
    ws.auto_filter.ref = "A3:G3"

    # merge row before footer
    merge_start_row = ws.max_row + 1
    ws.cell(row=merge_start_row, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.merge_cells(start_row=merge_start_row, start_column=1,
                   end_row=merge_start_row, end_column=7)
    ws.cell(row=merge_start_row, column=1)

    # increase height of merged cells
    ws.row_dimensions[merge_start_row].height = 20

    # write footer
    footer_row = ws.max_row + 1

    # footer row 1
    # 1st column
    ws.cell(row=footer_row, column=1, value="Telephone").font = Font(italic=True,
                                                                     size=11, bold=True)
    ws.cell(row=footer_row, column=1).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=1).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=1,
                   end_row=footer_row, end_column=2)
    # 2nd column
    ws.cell(row=footer_row, column=3, value="Address").font = Font(italic=True,
                                                                   size=11, bold=True)
    ws.cell(row=footer_row, column=3).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=3).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=3,
                   end_row=footer_row, end_column=5)
    # 3rd column
    ws.cell(row=footer_row, column=6, value="Email").font = Font(italic=True,
                                                                 size=11, bold=True)
    ws.cell(row=footer_row, column=6).fill = PatternFill(
        "solid", fgColor="ADD8E6")
    ws.cell(row=footer_row, column=6).alignment = Alignment(
        horizontal="center", vertical="top")
    ws.merge_cells(start_row=footer_row, start_column=6,
                   end_row=footer_row, end_column=7)

    ws.row_dimensions[footer_row].height = 20

    # footer row 2
    # 1st column
    ws.cell(row=footer_row + 1, column=1,
            value=group.telephone).font = Font(size=11)
    ws.cell(row=footer_row+1, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=1).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=1,
                   end_row=footer_row + 1, end_column=2)
    # 2nd column
    ws.cell(row=footer_row + 1, column=3,
            value=group.address).font = Font(size=11)
    ws.cell(row=footer_row+1, column=3).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=3).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=3,
                   end_row=footer_row + 1, end_column=5)

    # 3rd column
    ws.cell(row=footer_row + 1, column=6,
            value=group.email).font = Font(size=11)
    ws.cell(row=footer_row+1, column=6).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 1, column=6).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 1, start_column=6,
                   end_row=footer_row + 1, end_column=7)

    ws.row_dimensions[footer_row+1].height = 50

    # footer row 3 system name
    system_name = f"Report was generated by RCSG MIS on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.cell(row=footer_row + 2, column=1,
            value=system_name).font = Font(italic=True, color="800000", size=10)
    ws.cell(row=footer_row + 2, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 2, column=1).alignment = Alignment(wrap_text=True,
                                                                horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 2, start_column=1,
                   end_column=7, end_row=footer_row + 2)

    # footer row 4 copyrights
    ws.cell(row=footer_row + 3, column=1,
            value="© 2023 RCSG MIS developers").font = Font(italic=True, color="800000", size=10)
    ws.cell(row=footer_row + 3, column=1).fill = PatternFill(
        "solid", fgColor="D3D3D3")
    ws.cell(row=footer_row + 3, column=1).alignment = Alignment(wrap_text=True,
                                                                horizontal="center", vertical="center")
    ws.merge_cells(start_row=footer_row + 3, start_column=1,
                   end_column=7, end_row=footer_row + 3)

    # set print area
    ws.print_area = "A1:G" + str(ws.max_row)

    # set page orientation to landscape

    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # create an HttpResponse object with the Excel file as an attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=lend data - {dt} .xlsx'
    wb.save(response)

    return response
